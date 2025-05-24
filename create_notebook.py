import nbformat as nbf

# Create a new notebook
nb = nbf.v4.new_notebook()

# Add cells
cells = [
    nbf.v4.new_markdown_cell("""# Bitcoin MACD Trading Strategy Analysis

This notebook analyzes Bitcoin trading using MACD strategy. It includes:
1. Data loading and preprocessing
2. MACD signal generation
3. Backtesting
4. Performance analysis
5. Visualization"""),
    
    nbf.v4.new_code_cell("""# Import required libraries
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import os

# Set plot style
plt.style.use('seaborn')
sns.set_palette('husl')

# Create necessary directories
os.makedirs('data/plots', exist_ok=True)
os.makedirs('data/backtest_results', exist_ok=True)
os.makedirs('data/trade_analysis', exist_ok=True)
os.makedirs('data/weekly', exist_ok=True)"""),
    
    nbf.v4.new_markdown_cell("## 1. Data Loading and Preprocessing"),
    
    nbf.v4.new_code_cell("""# Load the hourly Bitcoin data
df_hourly = pd.read_csv('bitcoin_data_hourly.csv')
df_hourly['time'] = pd.to_datetime(df_hourly['Timestamp'])
df_hourly.set_index('time', inplace=True)

# Display first few rows
print("Hourly Data Sample:")
df_hourly.head()"""),
    
    nbf.v4.new_code_cell("""# Resample to weekly data
weekly_ohlc = df_hourly.resample('W').agg({
    'Open': 'first',
    'High': 'max',
    'Low': 'min',
    'Close': 'last',
    'Volume': 'sum'
})
weekly_ohlc.reset_index(inplace=True)

# Filter data from 2012 onwards
weekly_ohlc = weekly_ohlc[weekly_ohlc['time'] >= '2012-01-01']

# Display first few rows
print("Weekly Data Sample:")
weekly_ohlc.head()"""),
    
    nbf.v4.new_markdown_cell("## 2. Calculate Technical Indicators"),
    
    nbf.v4.new_code_cell("""# Calculate weekly moving averages
weekly_ohlc['fast_ema'] = weekly_ohlc['Close'].ewm(span=12, adjust=False).mean()
weekly_ohlc['slow_ema'] = weekly_ohlc['Close'].ewm(span=26, adjust=False).mean()
weekly_ohlc['sma_40'] = weekly_ohlc['Close'].rolling(window=40).mean()

# Calculate weekly MACD
weekly_ohlc['macd'] = weekly_ohlc['fast_ema'] - weekly_ohlc['slow_ema']
weekly_ohlc['macd_signal'] = weekly_ohlc['macd'].ewm(span=9, adjust=False).mean()
weekly_ohlc['macd_histogram'] = weekly_ohlc['macd'] - weekly_ohlc['macd_signal']

# Calculate MACD differences
weekly_ohlc['macd_diff'] = weekly_ohlc['macd'].diff()
weekly_ohlc['macd_histogram_diff'] = weekly_ohlc['macd_histogram'].diff()

# Identify bullish and bearish sides
weekly_ohlc['bullish_side'] = (weekly_ohlc['fast_ema'] > weekly_ohlc['slow_ema']).astype(int)
weekly_ohlc['bearish_side'] = (weekly_ohlc['fast_ema'] <= weekly_ohlc['slow_ema']).astype(int)
weekly_ohlc['below_sma_40'] = (weekly_ohlc['Close'] < weekly_ohlc['sma_40']).astype(int)

# Display indicators
weekly_ohlc[['Close', 'fast_ema', 'slow_ema', 'sma_40', 'macd', 'macd_signal', 'macd_histogram']].tail()"""),
    
    nbf.v4.new_markdown_cell("## 3. Generate Trading Signals"),
    
    nbf.v4.new_code_cell("""# Implement sell signal logic
weekly_ohlc['potential_sell_signal'] = ((weekly_ohlc['bullish_side'] == 1) & 
                                       (weekly_ohlc['macd_diff'] < 0)).astype(int)

# Implement buy signal logic 1
weekly_ohlc['macd_diff_positive'] = (weekly_ohlc['macd_diff'] > 0).astype(int)
weekly_ohlc['positive_count'] = 0
bearish_count = 0
positive_count_in_bearish = 0

for i in range(1, len(weekly_ohlc)):
    if weekly_ohlc.loc[weekly_ohlc.index[i], 'below_sma_40'] == 1:
        bearish_count += 1
        if weekly_ohlc.loc[weekly_ohlc.index[i], 'macd_diff_positive'] == 1:
            if bearish_count == 1:
                positive_count_in_bearish = 0
            positive_count_in_bearish += 1
            weekly_ohlc.loc[weekly_ohlc.index[i], 'positive_count'] = positive_count_in_bearish
    else:
        bearish_count = 0
        positive_count_in_bearish = 0

weekly_ohlc['potential_buy_signal_1'] = ((weekly_ohlc['below_sma_40'] == 1) & 
                                       (weekly_ohlc['positive_count'] == 2)).astype(int)

# Implement buy signal logic 2
weekly_ohlc['potential_buy_signal_2'] = ((weekly_ohlc['bullish_side'] == 1) & 
                                        (weekly_ohlc['macd_histogram_diff'] > 0)).astype(int)

# Combine both potential buy signals
weekly_ohlc['potential_buy_signal'] = ((weekly_ohlc['potential_buy_signal_1'] == 1) | 
                                      (weekly_ohlc['potential_buy_signal_2'] == 1)).astype(int)

# Display signals
signal_columns = ['Close', 'potential_buy_signal', 'potential_sell_signal', 
                 'potential_buy_signal_1', 'potential_buy_signal_2']
weekly_ohlc[signal_columns].tail()"""),
    
    nbf.v4.new_markdown_cell("## 4. Backtesting"),
    
    nbf.v4.new_code_cell("""# Initialize backtest parameters
initial_capital = 10000  # $10,000 USD
capital = initial_capital
btc_holdings = 0
positions = []
equity_curve = []
trades = []
position_type = "NONE"  # Track current position type: "LONG" or "NONE"

# Initialize actual signal columns
weekly_ohlc['buy_signal'] = 0
weekly_ohlc['sell_signal'] = 0

# Run through each week
for i in range(1, len(weekly_ohlc)):
    week = weekly_ohlc.iloc[i]
    prev_week = weekly_ohlc.iloc[i-1]
    
    # Calculate portfolio value at the start of the week
    if position_type == "LONG":
        portfolio_value = capital + btc_holdings * week['Open']
    else:  # NONE
        portfolio_value = capital
    
    # Check for buy signal
    if week['potential_buy_signal'] == 1 and position_type != "LONG":
        weekly_ohlc.loc[weekly_ohlc.index[i], 'buy_signal'] = 1
        
        if capital > 0:
            btc_bought = capital / week['Open']
            btc_holdings = btc_bought
            
            signal_type = "BUY (Bearish)" if week['potential_buy_signal_1'] == 1 else "BUY (Bullish)"
            
            trades.append({
                'date': week['time'],
                'type': signal_type,
                'price': week['Open'],
                'btc_amount': btc_bought,
                'usd_value': capital,
                'portfolio_value': portfolio_value
            })
            
            capital = 0
            position_type = "LONG"
    
    # Check for sell signal
    elif week['potential_sell_signal'] == 1 and position_type == "LONG":
        weekly_ohlc.loc[weekly_ohlc.index[i], 'sell_signal'] = 1
        
        if btc_holdings > 0:
            long_profit = btc_holdings * (week['Open'] - prev_week['Close'])
            capital += btc_holdings * week['Open']
            
            trades.append({
                'date': week['time'],
                'type': 'SELL',
                'price': week['Open'],
                'btc_amount': btc_holdings,
                'usd_value': btc_holdings * week['Open'],
                'portfolio_value': portfolio_value,
                'profit_loss': long_profit
            })
            
            btc_holdings = 0
            position_type = "NONE"
    
    # Calculate end-of-week portfolio value
    if position_type == "LONG":
        end_portfolio_value = capital + btc_holdings * week['Close']
    else:  # NONE
        end_portfolio_value = capital
    
    # Record position and portfolio value
    positions.append({
        'date': week['time'],
        'capital': capital,
        'btc_holdings': btc_holdings,
        'position_type': position_type,
        'btc_price': week['Close'],
        'portfolio_value': end_portfolio_value
    })
    
    equity_curve.append(end_portfolio_value)

# Convert positions and trades to DataFrames
positions_df = pd.DataFrame(positions)
trades_df = pd.DataFrame(trades) if trades else pd.DataFrame(columns=['date', 'type', 'price', 'btc_amount', 'usd_value', 'portfolio_value'])

# Display trade summary
print("Trade Summary:")
print(f"Total Trades: {len(trades_df)}")
print(f"Buy Signals: {len(trades_df[trades_df['type'].str.contains('BUY')])}")
print(f"Sell Signals: {len(trades_df[trades_df['type'] == 'SELL'])}")
print(f"Bearish Buy Signals: {len(trades_df[trades_df['type'] == 'BUY (Bearish)'])}")
print(f"Bullish Buy Signals: {len(trades_df[trades_df['type'] == 'BUY (Bullish)'])}")"""),
    
    nbf.v4.new_markdown_cell("## 5. Performance Analysis"),
    
    nbf.v4.new_code_cell("""# Calculate buy-and-hold performance
buy_and_hold_btc = initial_capital / weekly_ohlc.iloc[1]['Open']
buy_and_hold_value = buy_and_hold_btc * weekly_ohlc.iloc[-1]['Close']
buy_and_hold_return = (buy_and_hold_value / initial_capital - 1) * 100

# Calculate strategy performance metrics
final_portfolio_value = positions_df.iloc[-1]['portfolio_value'] if not positions_df.empty else 0
total_return = (final_portfolio_value / initial_capital - 1) * 100
max_drawdown = 0
peak = positions_df['portfolio_value'].iloc[0] if not positions_df.empty else 0

for value in positions_df['portfolio_value'] if not positions_df.empty else []:
    if value > peak:
        peak = value
    drawdown = (peak - value) / peak * 100
    if drawdown > max_drawdown:
        max_drawdown = drawdown

# Calculate additional metrics
num_trades = len(trades_df)
num_winning_trades = 0
total_profit = 0
total_loss = 0

if 'profit_loss' in trades_df.columns and not trades_df.empty:
    profitable_trades = trades_df[trades_df['profit_loss'] > 0]
    num_winning_trades = len(profitable_trades)
    total_profit = profitable_trades['profit_loss'].sum() if not profitable_trades.empty else 0
    
    losing_trades = trades_df[trades_df['profit_loss'] < 0]
    total_loss = abs(losing_trades['profit_loss'].sum()) if not losing_trades.empty else 0

win_rate = (num_winning_trades / len(trades_df[trades_df['type'] == 'SELL']) * 100) if len(trades_df[trades_df['type'] == 'SELL']) > 0 else 0
profit_factor = (total_profit / total_loss) if total_loss > 0 else float('inf')

# Calculate Sharpe ratio
positions_df['weekly_return'] = positions_df['portfolio_value'].pct_change()
strategy_returns = positions_df['weekly_return'].dropna()
strategy_mean_return = strategy_returns.mean()
strategy_std_return = strategy_returns.std()
strategy_sharpe = (strategy_mean_return / strategy_std_return) * np.sqrt(52)  # Annualized for weekly data

# Calculate buy-and-hold Sharpe ratio
weekly_ohlc['buy_hold_value'] = buy_and_hold_btc * weekly_ohlc['Close']
weekly_ohlc['buy_hold_return'] = weekly_ohlc['buy_hold_value'].pct_change()
buy_hold_returns = weekly_ohlc['buy_hold_return'].dropna()
buy_hold_mean_return = buy_hold_returns.mean()
buy_hold_std_return = buy_hold_returns.std()
buy_hold_sharpe = (buy_hold_mean_return / buy_hold_std_return) * np.sqrt(52)  # Annualized for weekly data

# Display performance metrics
print("Performance Metrics:")
print(f"Total Return: {total_return:.2f}%")
print(f"Buy-and-Hold Return: {buy_and_hold_return:.2f}%")
print(f"Maximum Drawdown: {max_drawdown:.2f}%")
print(f"Number of Trades: {num_trades}")
print(f"Win Rate: {win_rate:.2f}%")
print(f"Profit Factor: {profit_factor:.2f}")
print(f"Strategy Sharpe Ratio: {strategy_sharpe:.4f}")
print(f"Buy-and-Hold Sharpe Ratio: {buy_hold_sharpe:.4f}")
print(f"Sharpe Ratio Difference: {strategy_sharpe - buy_hold_sharpe:.4f}")"""),
    
    nbf.v4.new_markdown_cell("## 6. Visualizations"),
    
    nbf.v4.new_code_cell("""# Plot 1: Bitcoin price with 40-week SMA and signals (log scale)
plt.figure(figsize=(15, 10))
plt.subplot(2, 1, 1)
plt.title('Bitcoin Weekly Price with 40-week SMA and Signals (Long-Only, 2012-Present)')
plt.plot(weekly_ohlc['time'], weekly_ohlc['Close'], label='BTC Price', color='blue')
plt.plot(weekly_ohlc['time'], weekly_ohlc['sma_40'], label='40-week SMA', color='orange')

# Plot buy signals
bearish_buy_signals = weekly_ohlc[(weekly_ohlc['buy_signal'] == 1) & (weekly_ohlc['potential_buy_signal_1'] == 1)]
plt.scatter(bearish_buy_signals['time'], bearish_buy_signals['Close'], color='green', label='Buy Signal (Bearish)', marker='^', s=100)

bullish_buy_signals = weekly_ohlc[(weekly_ohlc['buy_signal'] == 1) & (weekly_ohlc['potential_buy_signal_2'] == 1) & (weekly_ohlc['potential_buy_signal_1'] == 0)]
plt.scatter(bullish_buy_signals['time'], bullish_buy_signals['Close'], color='lime', label='Buy Signal (Bullish)', marker='^', s=100)

# Plot sell signals
sell_signals = weekly_ohlc[weekly_ohlc['sell_signal'] == 1]
plt.scatter(sell_signals['time'], sell_signals['Close'], color='red', label='Sell Signal', marker='v', s=100)

plt.xlabel('Date')
plt.ylabel('Price (USD)')
plt.yscale('log')  # Set y-axis to logarithmic scale
plt.legend()
plt.grid(True)

# Plot 2: MACD and Signal Line with log scale
plt.subplot(2, 1, 2)
plt.title('MACD, Signal Line, and Histogram (Log Scale)')

# Filter out zero or negative values for log scale
macd_for_log = weekly_ohlc['macd'].copy()
macd_signal_for_log = weekly_ohlc['macd_signal'].copy()

# Add a small constant to make all values positive for log scale
min_value = min(macd_for_log.min(), macd_signal_for_log.min())
if min_value <= 0:
    offset = abs(min_value) + 1  # Add 1 to ensure all values are positive
    macd_for_log = macd_for_log + offset
    macd_signal_for_log = macd_signal_for_log + offset
    
    # Plot with log scale
    plt.plot(weekly_ohlc['time'], macd_for_log, label=f'MACD (offset +{offset:.2f})', color='blue')
    plt.plot(weekly_ohlc['time'], macd_signal_for_log, label=f'Signal Line (offset +{offset:.2f})', color='red')
    
    # Plot MACD histogram as bars (can't use log scale for this)
    plt.bar(weekly_ohlc['time'], weekly_ohlc['macd_histogram'], label='MACD Histogram', color='gray', alpha=0.3)
    
    # Highlight positive MACD difference in bearish territory
    positive_diff_bearish = weekly_ohlc[(weekly_ohlc['below_sma_40'] == 1) & (weekly_ohlc['macd_diff'] > 0)]
    if not positive_diff_bearish.empty:
        plt.scatter(positive_diff_bearish['time'], positive_diff_bearish['macd'] + offset, 
                   color='green', label='Positive MACD Diff in Bearish', marker='o', alpha=0.5)
    
    # Highlight positive MACD histogram difference in bullish territory
    positive_hist_diff_bullish = weekly_ohlc[(weekly_ohlc['bullish_side'] == 1) & (weekly_ohlc['macd_histogram_diff'] > 0)]
    if not positive_hist_diff_bullish.empty:
        plt.scatter(positive_hist_diff_bullish['time'], positive_hist_diff_bullish['macd'] + offset, 
                   color='lime', label='Positive Histogram Diff in Bullish', marker='o', alpha=0.5)
else:
    # If all values are already positive, no offset needed
    plt.plot(weekly_ohlc['time'], weekly_ohlc['macd'], label='MACD', color='blue')
    plt.plot(weekly_ohlc['time'], weekly_ohlc['macd_signal'], label='Signal Line', color='red')
    
    # Plot MACD histogram as bars
    plt.bar(weekly_ohlc['time'], weekly_ohlc['macd_histogram'], label='MACD Histogram', color='gray', alpha=0.3)
    
    # Highlight positive MACD difference in bearish territory
    positive_diff_bearish = weekly_ohlc[(weekly_ohlc['below_sma_40'] == 1) & (weekly_ohlc['macd_diff'] > 0)]
    plt.scatter(positive_diff_bearish['time'], positive_diff_bearish['macd'], 
               color='green', label='Positive MACD Diff in Bearish', marker='o', alpha=0.5)
    
    # Highlight positive MACD histogram difference in bullish territory
    positive_hist_diff_bullish = weekly_ohlc[(weekly_ohlc['bullish_side'] == 1) & (weekly_ohlc['macd_histogram_diff'] > 0)]
    plt.scatter(positive_hist_diff_bullish['time'], positive_hist_diff_bullish['macd'], 
               color='lime', label='Positive Histogram Diff in Bullish', marker='o', alpha=0.5)

plt.xlabel('Date')
plt.ylabel('MACD Value (Log Scale)')
plt.yscale('log')  # Set y-axis to logarithmic scale
plt.legend()
plt.grid(True)

plt.tight_layout()
plt.show()"""),
    
    nbf.v4.new_code_cell("""# Plot 3: Equity curve (log scale)
plt.figure(figsize=(15, 7))
plt.title('Strategy Equity Curve vs Buy-and-Hold (Long-Only, 2012-Present)')
plt.plot(positions_df['date'] if not positions_df.empty else [], positions_df['portfolio_value'] if not positions_df.empty else [], label='MACD Long-Only Strategy', color='blue')

# Calculate buy-and-hold equity curve
buy_hold_btc = initial_capital / weekly_ohlc.iloc[1]['Open']
buy_hold_equity = [buy_hold_btc * price for price in weekly_ohlc['Close'].iloc[1:]]
plt.plot(weekly_ohlc['time'].iloc[1:], buy_hold_equity, label='Buy and Hold', color='gray', linestyle='--')

plt.xlabel('Date')
plt.ylabel('Portfolio Value (USD)')
plt.yscale('log')  # Set y-axis to logarithmic scale
plt.legend()
plt.grid(True)
plt.show()"""),
    
    nbf.v4.new_code_cell("""# Plot 4: Position type over time with BTC price (log scale)
plt.figure(figsize=(15, 7))
plt.title('Position Type Over Time (Long-Only, 2012-Present)')

# Create a numeric representation of position type for plotting
position_type_numeric = positions_df['position_type'].map({'NONE': 0, 'LONG': 1})

# Plot position type
plt.plot(positions_df['date'], position_type_numeric, label='Position Type', color='blue', drawstyle='steps-post')

# Add Bitcoin price on secondary y-axis
ax1 = plt.gca()
ax2 = ax1.twinx()
ax2.plot(weekly_ohlc['time'].iloc[1:], weekly_ohlc['Close'].iloc[1:], label='BTC Price', color='gray', alpha=0.5)
ax2.set_yscale('log')  # Set secondary y-axis to logarithmic scale

# Add horizontal lines for position types
plt.axhline(y=1, color='green', linestyle='--', alpha=0.3, label='Long')
plt.axhline(y=0, color='gray', linestyle='--', alpha=0.3, label='None')

# Set y-axis ticks and labels
ax1.set_yticks([0, 1])
ax1.set_yticklabels(['NONE', 'LONG'])

plt.xlabel('Date')
ax1.set_ylabel('Position Type')
ax2.set_ylabel('BTC Price (USD)')

# Combine legends from both axes
lines1, labels1 = ax1.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')

plt.grid(True)
plt.show()"""),
    
    nbf.v4.new_code_cell("""# Plot 5: Sharpe ratio comparison
plt.figure(figsize=(10, 6))
plt.title('Sharpe Ratio Comparison: Long-Only vs Buy-and-Hold (2012-Present)')
strategies = ['MACD Long-Only', 'Buy-and-Hold']
sharpe_values = [strategy_sharpe, buy_hold_sharpe]

plt.bar(strategies, sharpe_values, color=['blue', 'gray'])
plt.ylabel('Sharpe Ratio')
plt.grid(axis='y')

# Add values on top of bars
for i, v in enumerate(sharpe_values):
    plt.text(i, v + 0.1, f"{v:.4f}", ha='center')

plt.tight_layout()
plt.show()"""),
    
    nbf.v4.new_markdown_cell("## 7. Save Results"),
    
    nbf.v4.new_code_cell("""# Save weekly dataset
weekly_data_path = 'data/weekly/BTC_USD_weekly_with_long_only_signals_from_2012.csv'
weekly_ohlc.to_csv(weekly_data_path, index=False)
print(f"Weekly dataset saved to: {weekly_data_path}")

# Save positions data
positions_df.to_csv('data/backtest_results/weekly_macd_long_only_from_2012_log_macd_positions.csv', index=False)
print("Positions data saved")

# Save trade data
trade_df = pd.DataFrame(trades)
if not trade_df.empty:
    trade_df = trade_df.sort_values('date')
    trade_csv_path = 'data/trade_analysis/weekly_macd_trades_from_2012_log_macd_with_signals.csv'
    trade_df.to_csv(trade_csv_path, index=False)
    print(f"Trade data saved to: {trade_csv_path}")
    
    # Create Excel version with formatting
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        trade_excel_path = 'data/trade_analysis/weekly_macd_trades_from_2012_log_macd_with_signals.xlsx'
        
        # Create a Pandas Excel writer
        writer = pd.ExcelWriter(trade_excel_path, engine='openpyxl')
        trade_df.to_excel(writer, index=False, sheet_name='MACD Trades')
        
        # Get the workbook and the worksheet
        workbook = writer.book
        worksheet = writer.sheets['MACD Trades']
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        sell_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply header styles
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Apply row styles based on action
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
            action_cell = worksheet.cell(row=row_idx, column=trade_df.columns.get_loc('type') + 1)
            
            # Apply fill based on action
            fill = buy_fill if 'BUY' in action_cell.value else sell_fill
            
            # Apply styles to all cells in the row
            for cell in row:
                cell.fill = fill
                cell.border = border
                
                # Center-align certain columns
                if cell.column in [trade_df.columns.get_loc('type') + 1, 
                                  trade_df.columns.get_loc('Below 40-week SMA') + 1,
                                  trade_df.columns.get_loc('Bullish Side') + 1]:
                    cell.alignment = Alignment(horizontal='center')
                
                # Right-align numeric columns
                if cell.column in [trade_df.columns.get_loc('price') + 1,
                                  trade_df.columns.get_loc('btc_amount') + 1,
                                  trade_df.columns.get_loc('usd_value') + 1,
                                  trade_df.columns.get_loc('portfolio_value') + 1]:
                    cell.alignment = Alignment(horizontal='right')
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the Excel file
        writer.close()
        print(f"Excel spreadsheet saved to: {trade_excel_path}")
        
    except ImportError:
        print("openpyxl not available, skipping Excel formatting")""")
]

# Add cells to notebook
nb['cells'] = cells

# Write notebook to file
with open('trading_analysis.ipynb', 'w', encoding='utf-8') as f:
    nbf.write(nb, f) 