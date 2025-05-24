import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
import os

def create_plots(weekly_ohlc, positions_df, trades_df, initial_capital, strategy_sharpe, buy_hold_sharpe, total_return, buy_and_hold_return, max_drawdown, num_trades, win_rate, profit_factor):
    """Create and save all visualization plots"""
    
    # Create directories if they don't exist
    os.makedirs('data/plots', exist_ok=True)
    os.makedirs('data/backtest_results', exist_ok=True)
    os.makedirs('data/trade_analysis', exist_ok=True)
    
    # Plot 1: MACD with log scale
    plt.figure(figsize=(15, 7))
    plt.title('MACD Analysis (Log Scale)')
    
    # Check if we need to offset MACD values for log scale
    min_macd = min(weekly_ohlc['macd'].min(), weekly_ohlc['macd_signal'].min())
    if min_macd <= 0:
        offset = abs(min_macd) + 1
        macd_for_log = weekly_ohlc['macd'] + offset
        macd_signal_for_log = weekly_ohlc['macd_signal'] + offset
        
        plt.plot(weekly_ohlc['time'], macd_for_log, label=f'MACD (offset +{offset:.2f})', color='blue')
        plt.plot(weekly_ohlc['time'], macd_signal_for_log, label=f'Signal Line (offset +{offset:.2f})', color='red')
    else:
        plt.plot(weekly_ohlc['time'], weekly_ohlc['macd'], label='MACD', color='blue')
        plt.plot(weekly_ohlc['time'], weekly_ohlc['macd_signal'], label='Signal Line', color='red')
    
    # Plot MACD histogram as bars
    plt.bar(weekly_ohlc['time'], weekly_ohlc['macd_histogram'], label='MACD Histogram', color='gray', alpha=0.3)
    
    # Highlight positive MACD difference in bearish territory
    positive_diff_bearish = weekly_ohlc[(weekly_ohlc['below_sma_40'] == 1) & (weekly_ohlc['macd_diff'] > 0)]
    if not positive_diff_bearish.empty:
        plt.scatter(positive_diff_bearish['time'], positive_diff_bearish['macd'] + (offset if min_macd <= 0 else 0), 
                   color='green', label='Positive MACD Diff in Bearish', marker='o', alpha=0.5)
    
    # Highlight positive MACD histogram difference in bullish territory
    positive_hist_diff_bullish = weekly_ohlc[(weekly_ohlc['bullish_side'] == 1) & (weekly_ohlc['macd_histogram_diff'] > 0)]
    if not positive_hist_diff_bullish.empty:
        plt.scatter(positive_hist_diff_bullish['time'], positive_hist_diff_bullish['macd'] + (offset if min_macd <= 0 else 0), 
                   color='lime', label='Positive Histogram Diff in Bullish', marker='o', alpha=0.5)
    
    plt.xlabel('Date')
    plt.ylabel('MACD Value (Log Scale)')
    plt.yscale('log')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig('data/plots/weekly_macd_long_only_from_2012_log_macd_signals.png')
    plt.close()
    
    # Plot 2: Equity curve (log scale)
    plt.figure(figsize=(15, 7))
    plt.title('Strategy Equity Curve vs Buy-and-Hold (Long-Only, 2012-Present)')
    plt.plot(positions_df['date'] if not positions_df.empty else [], 
             positions_df['portfolio_value'] if not positions_df.empty else [], 
             label='MACD Long-Only Strategy', color='blue')
    
    # Calculate buy-and-hold equity curve
    buy_hold_btc = initial_capital / weekly_ohlc.iloc[1]['open']
    buy_hold_equity = [buy_hold_btc * price for price in weekly_ohlc['close'].iloc[1:]]
    plt.plot(weekly_ohlc['time'].iloc[1:], buy_hold_equity, label='Buy and Hold', color='gray', linestyle='--')
    
    plt.xlabel('Date')
    plt.ylabel('Portfolio Value (USD)')
    plt.yscale('log')
    plt.legend()
    plt.grid(True)
    plt.savefig('data/plots/weekly_macd_long_only_from_2012_log_macd_equity_curve.png')
    plt.close()
    
    # Plot 3: Position type over time with BTC price (log scale)
    plt.figure(figsize=(15, 7))
    plt.title('Position Type Over Time (Long-Only, 2012-Present)')
    
    # Create a numeric representation of position type for plotting
    position_type_numeric = positions_df['position_type'].map({'NONE': 0, 'LONG': 1})
    
    # Plot position type
    plt.plot(positions_df['date'], position_type_numeric, label='Position Type', color='blue', drawstyle='steps-post')
    
    # Add Bitcoin price on secondary y-axis
    ax1 = plt.gca()
    ax2 = ax1.twinx()
    ax2.plot(weekly_ohlc['time'].iloc[1:], weekly_ohlc['close'].iloc[1:], label='BTC Price', color='gray', alpha=0.5)
    ax2.set_yscale('log')
    
    # Add horizontal lines for position types
    plt.axhline(y=1, color='green', linestyle='--', alpha=0.3, label='Long')
    plt.axhline(y=0, color='gray', linestyle='--', alpha=0.3, label='None')
    
    # Set y-axis ticks and labels
    ax1.set_yticks([0, 1])
    ax1.set_yticklabels(['NONE', 'LONG'])
    
    plt.xlabel('Date')
    ax1.set_ylabel('Position Type')
    ax2.set_ylabel('BTC Price (USD)')
    
    # Combine legends from both axes
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
    
    plt.grid(True)
    plt.savefig('data/plots/weekly_macd_long_only_from_2012_log_macd_position_type.png')
    plt.close()
    
    # Plot 4: Sharpe ratio comparison
    plt.figure(figsize=(10, 6))
    plt.title('Sharpe Ratio Comparison: Long-Only vs Buy-and-Hold (2012-Present)')
    strategies = ['MACD Long-Only', 'Buy-and-Hold']
    sharpe_values = [strategy_sharpe, buy_hold_sharpe]
    
    plt.bar(strategies, sharpe_values, color=['blue', 'gray'])
    plt.ylabel('Sharpe Ratio')
    plt.grid(axis='y')
    
    # Add values on top of bars
    for i, v in enumerate(sharpe_values):
        plt.text(i, v + 0.1, f"{v:.4f}", ha='center')
    
    plt.tight_layout()
    plt.savefig('data/plots/weekly_macd_long_only_from_2012_log_macd_sharpe.png')
    plt.close()
    
    # Save positions data
    positions_df.to_csv('data/backtest_results/weekly_macd_long_only_from_2012_log_macd_positions.csv', index=False)
    
    # Create trade spreadsheet
    print("Creating trade spreadsheet...")
    trade_data = []
    
    # Process each buy signal
    for i, row in weekly_ohlc[weekly_ohlc['buy_signal'] == 1].iterrows():
        # Determine signal type
        if row['potential_buy_signal_1'] == 1:
            signal_type = "BUY (Bearish)"
            signal_logic = "Price below 40-week SMA and MACD difference turned positive for the second time"
        else:
            signal_type = "BUY (Bullish)"
            signal_logic = "On bullish side (fast EMA > slow EMA) and MACD histogram difference turned positive"
        
        # Find matching trade in trades_df
        matching_trade = trades_df[trades_df['date'] == row['time']]
        
        if not matching_trade.empty:
            trade = matching_trade.iloc[0]
            trade_data.append({
                'Date': row['time'],
                'Action': signal_type,
                'Bitcoin Price': row['open'],
                'MACD': row['macd'],
                'MACD Signal Line': row['macd_signal'],
                'MACD Histogram': row['macd_histogram'],
                'MACD Diff': row['macd_diff'],
                'MACD Histogram Diff': row['macd_histogram_diff'],
                'Fast EMA': row['fast_ema'],
                'Slow EMA': row['slow_ema'],
                'Below 40-week SMA': 'Yes' if row['below_sma_40'] == 1 else 'No',
                'Bullish Side': 'Yes' if row['bullish_side'] == 1 else 'No',
                'Signal Logic': signal_logic,
                'Portfolio Value': trade['portfolio_value'],
                'BTC Amount': trade['btc_amount'],
                'USD Value': trade['usd_value']
            })
    
    # Process each sell signal
    for i, row in weekly_ohlc[weekly_ohlc['sell_signal'] == 1].iterrows():
        signal_logic = "On bullish side (fast EMA > slow EMA) and MACD difference turned negative"
        
        matching_trade = trades_df[trades_df['date'] == row['time']]
        
        if not matching_trade.empty:
            trade = matching_trade.iloc[0]
            
            # Calculate trade return
            if len(trade_data) > 0 and trade_data[-1]['Action'].startswith('BUY'):
                buy_price = trade_data[-1]['Bitcoin Price']
                sell_price = row['open']
                trade_return = (sell_price / buy_price - 1) * 100
                trade_return_str = f"{trade_return:.2f}%"
            else:
                trade_return_str = None
            
            trade_data.append({
                'Date': row['time'],
                'Action': 'SELL',
                'Bitcoin Price': row['open'],
                'MACD': row['macd'],
                'MACD Signal Line': row['macd_signal'],
                'MACD Histogram': row['macd_histogram'],
                'MACD Diff': row['macd_diff'],
                'MACD Histogram Diff': row['macd_histogram_diff'],
                'Fast EMA': row['fast_ema'],
                'Slow EMA': row['slow_ema'],
                'Below 40-week SMA': 'Yes' if row['below_sma_40'] == 1 else 'No',
                'Bullish Side': 'Yes' if row['bullish_side'] == 1 else 'No',
                'Signal Logic': signal_logic,
                'Portfolio Value': trade['portfolio_value'],
                'BTC Amount': trade['btc_amount'],
                'USD Value': trade['usd_value'],
                'Trade Return': trade_return_str
            })
    
    # Convert to dataframe and sort by date
    trade_df = pd.DataFrame(trade_data)
    if not trade_df.empty:
        trade_df = trade_df.sort_values('Date')
        
        # Save the trade spreadsheet
        trade_csv_path = 'data/trade_analysis/weekly_macd_trades_from_2012_log_macd_with_signals.csv'
        trade_df.to_csv(trade_csv_path, index=False)
        
        # Create Excel version with formatting
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            print("Creating formatted Excel spreadsheet...")
            trade_excel_path = 'data/trade_analysis/weekly_macd_trades_from_2012_log_macd_with_signals.xlsx'
            
            writer = pd.ExcelWriter(trade_excel_path, engine='openpyxl')
            trade_df.to_excel(writer, index=False, sheet_name='MACD Trades')
            
            workbook = writer.book
            worksheet = writer.sheets['MACD Trades']
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            sell_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Apply header styles
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Apply row styles
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
                action_cell = worksheet.cell(row=row_idx, column=trade_df.columns.get_loc('Action') + 1)
                fill = buy_fill if 'BUY' in action_cell.value else sell_fill
                
                for cell in row:
                    cell.fill = fill
                    cell.border = border
                    
                    if cell.column in [trade_df.columns.get_loc('Action') + 1, 
                                     trade_df.columns.get_loc('Below 40-week SMA') + 1,
                                     trade_df.columns.get_loc('Bullish Side') + 1]:
                        cell.alignment = Alignment(horizontal='center')
                    
                    if cell.column in [trade_df.columns.get_loc('Bitcoin Price') + 1,
                                     trade_df.columns.get_loc('MACD') + 1,
                                     trade_df.columns.get_loc('MACD Signal Line') + 1,
                                     trade_df.columns.get_loc('MACD Histogram') + 1,
                                     trade_df.columns.get_loc('MACD Diff') + 1,
                                     trade_df.columns.get_loc('MACD Histogram Diff') + 1,
                                     trade_df.columns.get_loc('Fast EMA') + 1,
                                     trade_df.columns.get_loc('Slow EMA') + 1,
                                     trade_df.columns.get_loc('Portfolio Value') + 1,
                                     trade_df.columns.get_loc('BTC Amount') + 1,
                                     trade_df.columns.get_loc('USD Value') + 1]:
                        cell.alignment = Alignment(horizontal='right')
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            writer.close()
            print(f"Excel spreadsheet saved to: {trade_excel_path}")
            
        except ImportError:
            print("openpyxl not available, skipping Excel formatting")
        
        print(f"Trade spreadsheet saved to: {trade_csv_path}")
        
        # Print trade summary
        print(f"\nTrade Summary:")
        print(f"Total Trades: {len(trade_df)}")
        print(f"Buy Signals: {len(trade_df[trade_df['Action'].str.contains('BUY')])}")
        print(f"Sell Signals: {len(trade_df[trade_df['Action'] == 'SELL'])}")
        print(f"Bearish Buy Signals: {len(trade_df[trade_df['Action'] == 'BUY (Bearish)'])}")
        print(f"Bullish Buy Signals: {len(trade_df[trade_df['Action'] == 'BUY (Bullish)'])}")
    else:
        print("No trades were executed during the backtest period.")
    
    # Print key performance metrics
    print("\nKey Performance Metrics:")
    print(f"Total Return: {total_return:.2f}%")
    print(f"Buy-and-Hold Return: {buy_and_hold_return:.2f}%")
    print(f"Maximum Drawdown: {max_drawdown:.2f}%")
    print(f"Number of Trades: {num_trades}")
    print(f"Win Rate: {win_rate:.2f}%")
    print(f"Profit Factor: {profit_factor:.2f}")
    print(f"Strategy Sharpe Ratio: {strategy_sharpe:.4f}")
    print(f"Buy-and-Hold Sharpe Ratio: {buy_hold_sharpe:.4f}")
    print(f"Sharpe Ratio Difference: {strategy_sharpe - buy_hold_sharpe:.4f}")
    
    print("\nVisualization completed successfully!")
    print("All plots and analysis files have been saved to the data directory.")

def main():
    # Load the hourly Bitcoin data
    hourly_data_path = 'bitcoin_data_hourly.csv'

    # Load the weekly Bitcoin data
    weekly_data_path = 'data/bitcoin_data_weekly.csv'

    # Load the positions data
    positions_df_path = 'data/backtest_results/weekly_macd_long_only_from_2012_log_macd_positions.csv'

    # Load the trades data
    trades_df_path = 'data/trade_analysis/weekly_macd_trades_from_2012_log_macd_with_signals.csv'

    # Load the strategy metrics
    strategy_metrics_path = 'data/backtest_results/weekly_macd_long_only_from_2012_log_macd_strategy_metrics.csv'

    # Load the buy-and-hold metrics
    buy_and_hold_metrics_path = 'data/backtest_results/weekly_macd_long_only_from_2012_log_macd_buy_and_hold_metrics.csv'

    # Load the weekly OHLC data
    weekly_ohlc = pd.read_csv(weekly_data_path)

    # Load the positions data
    positions_df = pd.read_csv(positions_df_path)

    # Load the trades data
    trades_df = pd.read_csv(trades_df_path)

    # Load the strategy metrics
    strategy_metrics = pd.read_csv(strategy_metrics_path)

    # Load the buy-and-hold metrics
    buy_and_hold_metrics = pd.read_csv(buy_and_hold_metrics_path)

    # Extract strategy metrics
    strategy_sharpe = strategy_metrics['sharpe_ratio'].values[0]
    buy_hold_sharpe = buy_and_hold_metrics['sharpe_ratio'].values[0]
    total_return = strategy_metrics['total_return'].values[0]
    buy_and_hold_return = buy_and_hold_metrics['total_return'].values[0]
    max_drawdown = strategy_metrics['max_drawdown'].values[0]
    num_trades = strategy_metrics['num_trades'].values[0]
    win_rate = strategy_metrics['win_rate'].values[0]
    profit_factor = strategy_metrics['profit_factor'].values[0]

    create_plots(weekly_ohlc, positions_df, trades_df, 1000, strategy_sharpe, buy_hold_sharpe, total_return, buy_and_hold_return, max_drawdown, num_trades, win_rate, profit_factor) 